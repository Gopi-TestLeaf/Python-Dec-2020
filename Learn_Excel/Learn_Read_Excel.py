# Step 1: install openpyxl - pip install openpyxl

# Read the data from Excel
import openpyxl

# Load the workbook(Excel file)
wb = openpyxl.load_workbook('C:\Python\Python_Dec\data\TC001.xlsx')

# returns all the sheet names of wb
print(wb.sheetnames)

# retuns active sheet name of wb
print(wb.active)

# create the object for specific sheet
sh = wb['Gopi']

# for row count
print("total rows count: ", sh.max_row)
print("total column count: ", sh.max_column)

# Way 1
# for row in range(1, sh.max_row+1):                      # for row
#     for column in range(1,sh.max_column+1):             # for column
#         print(sh.cell(row, column).value)

# read the value using index - 1, 1
# sh.cell(1,1).value


# Way2 2
for row in sh['A1' : 'E4']:
    for column in row:
        print(column.value)
    print("*"*25)

# read the value using excel add - A4
#print(sh['B2'].value)